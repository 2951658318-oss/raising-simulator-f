import openpyxl
import json

items = []

wb = openpyxl.load_workbook("物品与情感共鸣表.xlsx")

sheet = wb.active

for row in range(2, sheet.max_row + 1):

    item = {
        "id": sheet.cell(row,1).value,
        "image": "assets/items/" + sheet.cell(row,3).value,
        "filename": sheet.cell(row,3).value,
        "name": sheet.cell(row,4).value,
        "description": sheet.cell(row,5).value,
        "scores":{
            "collector": sheet.cell(row,6).value,
            "dreamer": sheet.cell(row,7).value,
            "observer": sheet.cell(row,8).value,
            "resonator": sheet.cell(row,9).value,
            "creator": sheet.cell(row,10).value,
            "thinker": sheet.cell(row,11).value,
            "player": sheet.cell(row,12).value,
            "performer": sheet.cell(row,13).value
        }
    }

    items.append(item)


with open("data/objects.json","w",encoding="utf-8") as f:
    json.dump(items,f,ensure_ascii=False,indent=2)


print("转换完成！共生成",len(items),"个物品")